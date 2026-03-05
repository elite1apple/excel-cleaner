"""
Excel Cleaning Script for Window Shade Automation
Cleans raw Excel files into standardized format for automation processing
"""

import pandas as pd
import re
from pathlib import Path
from typing import Dict, Optional, Tuple
import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter


def parse_fraction_string(value_str: str) -> Optional[float]:
    """
    Parse a string containing numbers with fractions, including Unicode fraction characters.
    
    Handles:
    - Regular decimals: "54", "51.5"
    - Mixed fractions with spaces: "54 3/4", "51 1/4"
    - Unicode fractions: "54¾", "51¼", "80⅛", "54½"
    - Regular fractions: "3/4", "1/4"
    
    Examples:
    - "54¾" → 54.75
    - "51 1/4" → 51.25
    - "80⅛" → 80.125
    - "51" → 51.0
    """
    if not value_str or pd.isna(value_str):
        return None
    
    # Unicode fraction mappings
    unicode_fractions = {
        '¼': 0.25,
        '½': 0.5,
        '¾': 0.75,
        '⅐': 1/7,
        '⅑': 1/9,
        '⅒': 0.1,
        '⅓': 1/3,
        '⅔': 2/3,
        '⅕': 0.2,
        '⅖': 0.4,
        '⅗': 0.6,
        '⅘': 0.8,
        '⅙': 1/6,
        '⅚': 5/6,
        '⅛': 0.125,
        '⅜': 0.375,
        '⅝': 0.625,
        '⅞': 0.875,
    }
    
    value_str = str(value_str).strip()
    
    # Fix common data entry typos
    value_str = re.sub(r'\.{2,}', '.', value_str)  # Double/triple dots "65..5" -> "65.5"
    value_str = value_str.replace(',', '.')         # Comma as decimal "62,5" -> "62.5"
    # Note: incomplete fractions like "62 1/" or "62 3/" are NOT auto-fixed because
    # they could be any 1/8-based fraction (1/8, 1/4, 3/8, 1/2, 5/8, 3/4, 7/8).
    
    # Replace Unicode fractions with their decimal equivalents
    for unicode_frac, decimal_val in unicode_fractions.items():
        if unicode_frac in value_str:
            # Extract whole number part if present
            parts = value_str.split(unicode_frac)
            if parts[0].strip():
                try:
                    whole = float(parts[0].strip())
                    return whole + decimal_val
                except (ValueError, TypeError):
                    pass
            else:
                return decimal_val
    
    # Handle regular fractions with spaces (e.g., "54 3/4" or "51 1/4")
    if ' ' in value_str and '/' in value_str:
        parts = value_str.split()
        if len(parts) == 2:
            try:
                whole = float(parts[0])
                frac_parts = parts[1].split('/')
                if len(frac_parts) == 2:
                    numerator = float(frac_parts[0])
                    denominator = float(frac_parts[1])
                    return whole + (numerator / denominator)
            except (ValueError, TypeError, ZeroDivisionError):
                pass
    
    # Handle standalone fractions (e.g., "3/4")
    if '/' in value_str and ' ' not in value_str:
        parts = value_str.split('/')
        if len(parts) == 2:
            try:
                numerator = float(parts[0])
                denominator = float(parts[1])
                return numerator / denominator
            except (ValueError, TypeError, ZeroDivisionError):
                pass
    
    # Try regular float parsing
    try:
        return float(value_str.rstrip('.'))
    except (ValueError, TypeError):
        return None


def extract_deduction_values_from_notes(df_raw: pd.DataFrame, header_row: int = 8, deduction_cell: str = None) -> Dict[str, float]:
    """
    Extract deduction values from notes section.
    
    First tries to read from a specific cell (e.g., "G8" or "I6"), then falls back to searching rows.
    
    Simple logic:
    - Look for "D = 1" or "D=1" in notes (can be fraction like "D=1/2" or decimal "D=1")
    - D value is the TOTAL value for both sides
    - Dl and Dr are automatically set to D/2 (half of D)
    
    Examples:
    - Notes: "D=1" → D=1.0, Dl=0.5, Dr=0.5
    - Notes: "D=1/2" → D=0.5, Dl=0.25, Dr=0.25
    
    Args:
        deduction_cell: Optional cell reference like "G8" or "I6" to read D value from
    
    Returns dict: {'D': value, 'Dl': D/2, 'Dr': D/2, 'DL': D/2, 'DR': D/2}
    """
    result = {}
    
    # First, try to read from specific cell if provided
    if deduction_cell:
        try:
            import re as _re
            from openpyxl.utils import column_index_from_string
            
            # Parse cell reference (e.g., "G8" -> col_letter="G", row_num=8)
            _m = _re.match(r'^([A-Za-z]+)(\d+)$', deduction_cell.strip())
            if not _m:
                raise ValueError(f"Invalid cell reference: {deduction_cell}")
            col_letter, row_num = _m.group(1).upper(), int(_m.group(2))
            col_idx = column_index_from_string(col_letter) - 1  # Convert to 0-based index
            row_idx = row_num - 1  # Convert to 0-based index
            
            # Check if row/col are within bounds
            if 0 <= row_idx < len(df_raw) and 0 <= col_idx < len(df_raw.columns):
                cell_value = df_raw.iloc[row_idx, col_idx]
                if pd.notna(cell_value):
                    cell_text = str(cell_value)
                    print(f"DEBUG: Reading from cell {deduction_cell}: {cell_text}")
                    
                    # Look for D=value pattern in the cell
                    d_pattern = re.search(r'D\s*[=:]\s*((?:\d+/\d+)|(?:\d*\.?\d+))', cell_text, re.IGNORECASE)
                    if d_pattern:
                        value_str = d_pattern.group(1)
                        print(f"DEBUG: Matched D pattern in cell {deduction_cell}, value_str='{value_str}'")
                        
                        # Parse value
                        if '/' in value_str:
                            value = int(value_str.split('/')[0]) / int(value_str.split('/')[1])
                        else:
                            value = float(value_str)
                        
                        result['D'] = value
                        half_value = value / 2
                        result['Dl'] = half_value
                        result['Dr'] = half_value
                        result['DL'] = half_value
                        result['DR'] = half_value
                        
                        print(f"DEBUG: Extracted from cell {deduction_cell} - D={result['D']}, Dl={result['Dl']}, Dr={result['Dr']}")
                        return result
        except Exception as e:
            print(f"DEBUG: Could not read from cell {deduction_cell}: {e}")
    
    # Fall back to searching rows (check header section first, then bottom)
    print(f"DEBUG extract_deduction_values_from_notes: Searching for D value in rows...")
    
    # Search header section first (rows near header_row)
    for idx in range(max(0, header_row - 5), min(header_row + 5, len(df_raw))):
        row_text = ' '.join([str(x) for x in df_raw.iloc[idx].tolist() if pd.notna(x)])
        
        # Look for D=value pattern
        d_pattern = re.search(r'D\s*[=:]\s*((?:\d+/\d+)|(?:\d*\.?\d+))', row_text, re.IGNORECASE)
        if d_pattern:
            value_str = d_pattern.group(1)
            print(f"DEBUG: Found D pattern in header row {idx}: {row_text[:200]}")
            print(f"DEBUG: Matched D pattern, value_str='{value_str}'")
            
            # Parse value
            if '/' in value_str:
                parts = value_str.split('/')
                value = int(parts[0]) / int(parts[1])
                print(f"DEBUG: Parsed as fraction: {value}")
            else:
                value = float(value_str)
                print(f"DEBUG: Parsed as decimal: {value}")
            
            result['D'] = value
            half_value = value / 2
            result['Dl'] = half_value
            result['Dr'] = half_value
            result['DL'] = half_value
            result['DR'] = half_value
            
            print(f"DEBUG: Final result - D={result['D']}, Dl={result['Dl']}, Dr={result['Dr']}")
            return result
    
    # If not found in header, search bottom rows (skip data rows)
    print(f"DEBUG: Not found in header, searching bottom rows...")
    for idx in range(len(df_raw) - 1, max(len(df_raw) - 20, header_row), -1):
        row_text = ' '.join([str(x) for x in df_raw.iloc[idx].tolist() if pd.notna(x)])
        
        # Skip if this looks like a data row
        first_col = df_raw.iloc[idx, 0] if len(df_raw.columns) > 0 else None
        if pd.notna(first_col):
            try:
                float(str(first_col).strip())
                continue  # Skip data rows
            except (ValueError, TypeError):
                pass
        
        # Look for D=value pattern
        d_pattern = re.search(r'D\s*[=:]\s*((?:\d+/\d+)|(?:\d*\.?\d+))', row_text, re.IGNORECASE)
        if d_pattern:
            value_str = d_pattern.group(1)
            print(f"DEBUG: Found D pattern in notes row {idx}: {row_text[:200]}")
            
            if '/' in value_str:
                value = int(value_str.split('/')[0]) / int(value_str.split('/')[1])
            else:
                value = float(value_str)
            
            result['D'] = value
            half_value = value / 2
            result['Dl'] = half_value
            result['Dr'] = half_value
            result['DL'] = half_value
            result['DR'] = half_value
            
            print(f"DEBUG: Final result - D={result['D']}, Dl={result['Dl']}, Dr={result['Dr']}")
            return result
    
    return result


def parse_deduction_value(deduct_code: str, numeric_value: float, notes_text: str = "", default_values: Dict[str, float] = None) -> Optional[float]:
    """
    Parse deduction value from Deducts code.
    
    Simple logic:
    - If Deducts = "D": use D value from notes (total value)
    - If Deducts = "Dl": use Dl value from notes (which is D/2)
    - If Deducts = "Dr": use Dr value from notes (which is D/2)
    
    If code has explicit value (e.g., "D=1" or "D-1/2"), parse and return that.
    Otherwise, use default_values from notes.
    """
    if pd.isna(deduct_code):
        return None
    
    deduct_code_str = str(deduct_code).strip()
    deduct_code_upper = deduct_code_str.upper()
    
    # Extract code part (D, Dl, or Dr) - handle formats like "D-1/2" or just "D"
    code_match = re.match(r'^(D[LR]?)\s*[-=]\s*((?:\d+/\d+)|(?:\d+\.?\d*))', deduct_code_upper)
    
    if code_match:
        # Code has explicit value (e.g., "D=1" or "D-1/2")
        code_part = code_match.group(1).upper()
        value_str = code_match.group(2)
        
        # Parse value
        if '/' in value_str:
            # Fraction (e.g., "1/2")
            parts = value_str.split('/')
            value = int(parts[0]) / int(parts[1])
        else:
            # Decimal (e.g., "1")
            value = float(value_str)
        
        # For D codes:
        # - "D=1" means total is 1.0, return as-is
        # - "D-1/2" means per-side is 1/2, total is 1.0, return 1.0
        if code_part == 'D':
            if '=' in deduct_code_str:
                return value  # D=1 → total is 1.0
            else:
                return value * 2  # D-1/2 → per-side is 1/2, total is 1.0
        else:
            # Dl or Dr: value is per-side, return as-is
            return value
    
    # No explicit value in code, check numeric_value column
    if not pd.isna(numeric_value):
        if deduct_code_upper == 'D':
            return float(numeric_value) * 2  # Total = 2 * per-side
        else:
            return float(numeric_value)  # Per-side value
    
    # No explicit value, use default_values from notes
    if default_values:
        if deduct_code_upper == 'D':
            # Return D value (total)
            d_value = default_values.get('D', 1.0)
            if 'D' not in default_values:
                print(f"DEBUG parse_deduction_value: 'D' not in default_values, using fallback 1.0")
            else:
                print(f"DEBUG parse_deduction_value: Found D={d_value} in default_values")
            return d_value
        elif deduct_code_upper in ['DL', 'Dl']:
            # Return Dl value (which is D/2) - check both cases
            return default_values.get('Dl', default_values.get('DL', 0.5))
        elif deduct_code_upper in ['DR', 'Dr']:
            # Return Dr value (which is D/2) - check both cases
            return default_values.get('Dr', default_values.get('DR', 0.5))
    
    # No default values, use fallback defaults
    if deduct_code_upper == 'D':
        return 1.0  # Default total
    else:
        return 0.5  # Default per-side


def normalize_fabric(fabric_str: str) -> str:
    """
    Normalize fabric name to a canonical form.
    Strips trailing numbers/spaces so variants like "Bed1", "Bed 2", "Bedroom",
    "Kitchen1", "Studio 2" are all reduced to their root canonical name.

    Known families (prefix → canonical):
        BED / BEDROOM  -> "Bed"
        LIV / LIVING   -> "Liv"
        KITCHEN        -> "Kitchen"
        STUDIO         -> "Studio"
        DEN            -> "Den"
        BATH / BATHROOM-> "Bath"
        LAUNDRY        -> "Laundry"
        OFFICE         -> "Office"
        DINING         -> "Dining"
        GUEST          -> "Guest"
        FAMILY         -> "Family"

    Everything else is returned as-is (stripped).
    """
    s = fabric_str.strip()
    upper = s.upper().replace(' ', '')  # collapse spaces for prefix matching

    FABRIC_FAMILIES = [
        ('BED',     'Bed'),
        ('LIV',     'Liv'),
        ('KITCHEN', 'Kitchen'),
        ('STUDIO',  'Studio'),
        ('DEN',     'Den'),
        ('BATH',    'Bath'),
        ('LAUNDRY', 'Laundry'),
        ('OFFICE',  'Office'),
        ('DINING',  'Dining'),
        ('GUEST',   'Guest'),
        ('FAMILY',  'Family'),
    ]

    # Capture any trailing digits at the very end of the string
    trailing_match = re.search(r'(\d+)$', upper)
    trailing_num = trailing_match.group(1) if trailing_match else ""

    for prefix, canonical in FABRIC_FAMILIES:
        if upper.startswith(prefix):
            return canonical + trailing_num

    # If no known prefix matched, try preserving trailing numbers on the original string
    # (By default it just returns the stripped original string anyway)
    return s


def parse_chain_value(chain_value) -> Optional[float]:
    """
    Parse chain value from raw file, handling both numeric and string forms.
    Accepts: 60, 60.0, "60", "60""
    Returns float, or None if unparseable.
    """
    if chain_value is None or (isinstance(chain_value, float) and pd.isna(chain_value)):
        return None
    # If already numeric
    if isinstance(chain_value, (int, float)):
        return float(chain_value)
    # String: strip whitespace, strip trailing inch mark
    s = str(chain_value).strip().rstrip('"').strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def format_control_length(chain_value) -> Optional[str]:
    """Format chain value as ControlLength (e.g., 72 or '72"' -> '72" LOOP')"""
    parsed = parse_chain_value(chain_value)
    if parsed is None:
        return None
    return f'{int(parsed)}" LOOP'


def create_special_instructions(
    has_reverse_roll: bool,
    control_side: str,
    extra_fabric_deduction: Optional[float],
    deduction_code: Optional[str] = None
) -> str:
    """Generate SpecialInstructions based on reverse roll and deductions
    
    Formatting matches image exactly (line by line):
    - When ReverseRoll=Yes and no deduction: "Apply Punch Reverse Fascia !!!"
    - When ReverseRoll=Yes and has deduction:
      Line 1: "Apply Punch Reverse Fascia !!!"
      Line 2: "--"
      Line 3: "Fabric Deduction 1/2" from LEFT SIDE ONLY!!!" (or "from both sides" for D code)
    - When ReverseRoll=No: Only deduction info if present, no Punch Reverse Fascia
    
    Args:
        deduction_code: "D" (both sides), "Dl" (left), "Dr" (right), or None
    """
    instructions = []
    
    # Determine side text based on deduction code (not control_side)
    # If deduction code is Dl/Dr, use that; otherwise use control_side
    if deduction_code:
        if deduction_code.upper() == 'D':
            side_text = "both sides"
        elif deduction_code.upper() == 'DL':
            side_text = "LEFT SIDE ONLY"
        elif deduction_code.upper() == 'DR':
            side_text = "RIGHT SIDE ONLY"
        else:
            # Fallback to control_side if deduction_code is unexpected
            side_text = "LEFT SIDE ONLY" if control_side == "LEFT" else "RIGHT SIDE ONLY"
    else:
        # No deduction code, use control_side
        side_text = "LEFT SIDE ONLY" if control_side == "LEFT" else "RIGHT SIDE ONLY"
    
    # Only add Punch Reverse Fascia when ReverseRoll is Yes
    if has_reverse_roll:
        # Add header line matching user request
        instructions.append(".---")
        instructions.append("Apply Punch Reverse Fascia !!!")
        
        # Add deduction info if present
        if extra_fabric_deduction is not None and extra_fabric_deduction > 0:
            # Format deduction (handle fractions)
            if extra_fabric_deduction == 0.5:
                deduction_text = '1/2"'
            elif extra_fabric_deduction == 0.25:
                deduction_text = '1/4"'
            elif extra_fabric_deduction == 1.0:
                deduction_text = '1"'
            else:
                deduction_text = f'{extra_fabric_deduction}"'
            
            # Add separator (3 dashes) and deduction line
            instructions.append("---")
            instructions.append(f"Fabric Deduction {deduction_text} from {side_text} !!!")
    else:
        # When ReverseRoll is No, only add deduction info if present
        if extra_fabric_deduction is not None and extra_fabric_deduction > 0:
            if extra_fabric_deduction == 0.5:
                deduction_text = '1/2"'
            elif extra_fabric_deduction == 0.25:
                deduction_text = '1/4"'
            elif extra_fabric_deduction == 1.0:
                deduction_text = '1"'
            else:
                deduction_text = f'{extra_fabric_deduction}"'
            
            # Just add the deduction line (no header/branding for now unless requested)
            instructions.append(f"Fabric Deduction {deduction_text} from {side_text} !!!")
    
    return "\n".join(instructions) if instructions else ""


def determine_reverse_roll(roll_value: str) -> Optional[str]:
    """Convert Roll value to ReverseRoll (Rev -> Yes)"""
    if pd.isna(roll_value):
        return None
    if str(roll_value).strip().upper() == "REV":
        return "Yes"
    return None


def extract_color_codes_from_header(df_raw: pd.DataFrame, header_row: int = 8) -> Dict[str, str]:
    """
    Extract color codes from header/metadata section (rows 0-7).
    
    Looks for patterns like:
    - "Bed = COLOR123" or "Bedroom = COLOR123"
    - "Living = COLOR456" or "Liv = COLOR456"
    - "Bed/Bedroom = COLOR123"
    - "Living/Liv = COLOR456"
    - "Studio = COLOR789"
    - "Den = COLOR012"
    - "Kitchen = COLOR345"
    - "Bath = COLOR678"
    
    Returns dict with extracted colors: {'Bed': 'COLOR123', 'Liv': 'COLOR456', 'Studio': 'COLOR789', ...}
    """
    color_codes = {}
    
    # Define all fabric types and their patterns
    fabric_patterns = {
        'Bed': [
            r'Bed\s*[/=]\s*Bedroom\s*=\s*([A-Z0-9]+)',
            r'Bedroom\s*=\s*([A-Z0-9]+)',
            r'Bed\s*=\s*([A-Z0-9]+)',
        ],
        'Liv': [
            r'Living\s*[/=]\s*Liv\s*=\s*([A-Z0-9]+)',
            r'Living\s*=\s*([A-Z0-9]+)',
            r'Liv\s*=\s*([A-Z0-9]+)',
        ],
        'Studio': [
            r'Studio\s*=\s*([A-Z0-9]+)',
        ],
        'Den': [
            r'Den\s*=\s*([A-Z0-9]+)',
        ],
        'Kitchen': [
            r'Kitchen\s*=\s*([A-Z0-9]+)',
        ],
        'Bath': [
            r'Bath\s*=\s*([A-Z0-9]+)',
            r'Bathroom\s*=\s*([A-Z0-9]+)',
        ],
    }
    
    # Search header rows (0 to header_row-1)
    for idx in range(header_row):
        row_text = ' '.join([str(x) for x in df_raw.iloc[idx].tolist() if pd.notna(x)])
        
        # Check each fabric type
        for fabric_type, patterns in fabric_patterns.items():
            # Skip if already found
            if fabric_type in color_codes:
                continue
                
            for pattern in patterns:
                match = re.search(pattern, row_text, re.IGNORECASE)
                if match:
                    color_codes[fabric_type] = match.group(1).strip()
                    break
    
    return color_codes


def read_excel_tolerant(file_path: str, sheet_name=0, header=None) -> pd.DataFrame:
    """
    Read an Excel file tolerantly.

    xlsx files are ZIP archives.  When openpyxl chokes on a corrupt
    xl/styles.xml (stylesheet), we copy the archive to a temp file with that
    entry stripped out — openpyxl then uses a blank default stylesheet and
    reads the data fine.
    """
    import zipfile
    import shutil
    import tempfile

    def _read(path):
        return pd.read_excel(path, sheet_name=sheet_name, header=header)

    # ── Fast path ────────────────────────────────────────────────────────────
    _primary_err = None
    try:
        return _read(file_path)
    except Exception as primary_err:
        err_str = str(primary_err).lower()
        is_stylesheet_err = any(k in err_str for k in ('stylesheet', 'invalid xml', 'workbook'))
        if not is_stylesheet_err:
            raise
        _primary_err = primary_err  # save before Python deletes it at end of except block

    # ── Slow path: strip the bad stylesheet from the ZIP ─────────────────────
    print(f"WARNING: Excel stylesheet is corrupt ({_primary_err}). "
          f"Rebuilding file without xl/styles.xml ...")

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    try:
        with zipfile.ZipFile(file_path, 'r') as zin, \
             zipfile.ZipFile(tmp.name, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == 'xl/styles.xml':
                    # Replace with a minimal valid stylesheet so openpyxl is happy
                    minimal_styles = (
                        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
                        '<fills count="2">'
                        '<fill><patternFill patternType="none"/></fill>'
                        '<fill><patternFill patternType="gray125"/></fill>'
                        '</fills>'
                        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
                        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
                        '<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
                        '</styleSheet>'
                    )
                    zout.writestr(item, minimal_styles)
                    print("  → Replaced corrupt xl/styles.xml with minimal stylesheet")
                else:
                    zout.writestr(item, zin.read(item.filename))

        return _read(tmp.name)

    except Exception as fallback_err:
        raise RuntimeError(
            f"Could not read Excel file even after stripping stylesheet. "
            f"Primary error: {_primary_err}. Fallback error: {fallback_err}"
        ) from _primary_err
    finally:
        try:
            import os
            os.remove(tmp.name)
        except Exception:
            pass


def scan_text_tags(
    input_file: str,
    header_row: int = 8,
    notes_start_keywords: list = None
) -> list:
    """
    Quick scan: return a list of rows whose Tag/Unit column contains
    non-numeric text (e.g. 'th102', 'garage').  Does NOT produce output.

    Returns a list of dicts with keys: tag, fabric, width, height.
    """
    if notes_start_keywords is None:
        notes_start_keywords = ["Total", "all Finshed", "Punch Reverse", "Deducts"]

    df = read_excel_tolerant(input_file, sheet_name=0, header=header_row)

    # Apply notes filter
    mask = pd.Series([True] * len(df))
    for col in df.columns:
        if df[col].dtype == 'object':
            for kw in notes_start_keywords:
                mask &= ~df[col].astype(str).str.contains(kw, case=False, na=False)
    df = df[mask].copy()

    # Find tag column
    tag_col = next(
        (c for c in df.columns if str(c).upper() in ['TAG', 'TAG/UNIT', 'UNIT']),
        None
    )
    if not tag_col:
        return []

    def _is_text_tag(x):
        return isinstance(x, str) and not pd.isna(x) and not str(x).replace('.', '', 1).isdigit()

    text_rows = df[df[tag_col].apply(_is_text_tag)]

    height_col = 'Height' if 'Height' in df.columns else ('Length' if 'Length' in df.columns else None)

    results = []
    for _, row in text_rows.iterrows():
        w = row.get('Width', None)
        h = row.get(height_col, None) if height_col else None
        # Auto-skip rows with NO measurements (section labels like 'nrm', 'NOTES')
        w_empty = w is None or (isinstance(w, float) and pd.isna(w))
        h_empty = h is None or (isinstance(h, float) and pd.isna(h))
        if w_empty and h_empty:
            continue
        results.append({
            'tag': str(row.get(tag_col, '')),
            'fabric': str(row.get('Fabric', '')),
            'width': w if not w_empty else '',
            'height': h if not h_empty else '',
        })
    return results


def clean_excel_file(
    input_file: str,
    output_file: Optional[str] = None,
    bed_color: Optional[str] = None,
    liv_color: Optional[str] = None,
    fabric_colors: Optional[Dict[str, str]] = None,
    header_row: int = 8,
    notes_start_keywords: list = None,
    deduction_cell: Optional[str] = None,
    tag_action: str = 'skip'
) -> pd.DataFrame:
    """
    Clean raw Excel file into standardized format.
    
    Parameters:
    -----------
    input_file : str
        Path to input Excel file
    output_file : str, optional
        Path to save cleaned file (if None, auto-generates name)
    bed_color : str, optional
        Color number for Bed fabric (if None, will try to detect or prompt)
    liv_color : str, optional
        Color number for Liv fabric (if None, will try to detect or prompt)
    fabric_colors : dict, optional
        Dictionary mapping fabric types to color numbers
        Example: {'Kitchen': 'KITCHEN_COLOR', 'Studio': 'STUDIO_COLOR', 'Den': 'DEN_COLOR'}
    header_row : int
        Row number (0-indexed) where data headers are located
    notes_start_keywords : list
        Keywords that indicate start of notes section (rows to exclude)
    
    Returns:
    --------
    pd.DataFrame : Cleaned dataframe
    """
    if notes_start_keywords is None:
        notes_start_keywords = ["Total", "all Finshed", "Punch Reverse", "Deducts"]
    
    # Read the Excel file (tolerant mode handles corrupt stylesheets)
    print(f"Reading file: {input_file}")
    df_raw = read_excel_tolerant(input_file, sheet_name=0, header=None)
    
    # Extract header row
    headers = df_raw.iloc[header_row].tolist()
    
    # Extract color codes from header section
    header_colors = extract_color_codes_from_header(df_raw, header_row)
    
    # Use header colors if not provided as parameters
    if not bed_color and 'Bed' in header_colors:
        bed_color = header_colors['Bed']
        print(f"Found Bed color in header: {bed_color}")
    if not liv_color and 'Liv' in header_colors:
        liv_color = header_colors['Liv']
        print(f"Found Liv color in header: {liv_color}")
    
    # Merge header colors into fabric_colors dict (header colors take precedence)
    if fabric_colors is None:
        fabric_colors = {}
    
    # Add all header colors to fabric_colors (except Bed/Liv which are handled separately)
    for fabric_type, color_code in header_colors.items():
        if fabric_type not in ['Bed', 'Liv']:
            fabric_colors[fabric_type] = color_code
            print(f"Found {fabric_type} color in header: {color_code}")
    
    # Extract deduction values from notes section before reading data
    # Try specific cell first (default to I6)
    if deduction_cell is None:
        deduction_cell = 'I6'  # Default location
    deduction_values = extract_deduction_values_from_notes(df_raw, header_row, deduction_cell=deduction_cell)
    print(f"DEBUG: Notes extraction returned: {deduction_values}")
    if deduction_values:
        print(f"DEBUG: Extracted deduction values from notes:")
        for key, value in deduction_values.items():
            print(f"  {key}: {value}")
    else:
        print("DEBUG: No deduction values found in notes - will use defaults")
    
    # Read data starting from header_row + 1
    df = read_excel_tolerant(input_file, sheet_name=0, header=header_row)
    
    # Remove rows that are notes (check for keywords in Tag or other columns)
    print("Removing note rows...")
    mask = pd.Series([True] * len(df))
    
    for col in df.columns:
        if df[col].dtype == 'object':
            for keyword in notes_start_keywords:
                mask &= ~df[col].astype(str).str.contains(keyword, case=False, na=False)
    
    # Also remove rows where Tag is text (not numeric) and not NaN
    # Handle different column names: "Tag", "Tag/Unit", "Tag/unit", or "unit" (case-insensitive)
    tag_col = None
    for col in df.columns:
        col_upper = str(col).upper()
        if col_upper in ['TAG', 'TAG/UNIT', 'UNIT']:
            tag_col = col
            break
    
    # Separate / handle rows with non-numeric tags based on tag_action
    # tag_action: 'skip' (default), 'keep' (use text as-is), 'extract' (pull number out)
    skipped_df = pd.DataFrame()
    if tag_col:
        # 1. Apply notes/garbage filter first
        df = df[mask].copy()

        # 2. Identify non-numeric tags in the cleaned set
        def _is_text_tag(x):
            return isinstance(x, str) and not pd.isna(x) and not str(x).replace('.', '', 1).isdigit()

        non_numeric_mask = df[tag_col].apply(_is_text_tag)
        text_tag_rows = df[non_numeric_mask].copy()

        # Auto-skip text-tag rows that have NO width and NO height
        # (they are section headers / labels like 'nrm', not real data rows)
        _height_col_scan = 'Height' if 'Height' in text_tag_rows.columns else ('Length' if 'Length' in text_tag_rows.columns else None)
        def _has_measurements(row):
            w = row.get('Width', None)
            h = row.get(_height_col_scan, None) if _height_col_scan else None
            return not (pd.isna(w) if w is None or (isinstance(w, float) and __import__('math').isnan(w)) else pd.isna(w)) \
                   or not (pd.isna(h) if h is None or (isinstance(h, float) and __import__('math').isnan(h)) else pd.isna(h))

        if len(text_tag_rows) > 0 and _height_col_scan:
            has_data_mask = text_tag_rows.apply(
                lambda row: not (pd.isna(row.get('Width', float('nan'))) and pd.isna(row.get(_height_col_scan, float('nan')))),
                axis=1
            )
            auto_skipped = text_tag_rows[~has_data_mask]
            text_tag_rows = text_tag_rows[has_data_mask]
            if len(auto_skipped) > 0:
                print(f"DEBUG: Auto-skipped {len(auto_skipped)} text-tag rows with no measurements (section labels)")
            # Update the non_numeric_mask to only cover rows with actual data
            non_numeric_mask = df[tag_col].apply(_is_text_tag)
            # Build a combined mask: text-tag rows that also HAVE measurements
            rows_with_data_idx = set(text_tag_rows.index)
            non_numeric_mask = non_numeric_mask & df.index.isin(rows_with_data_idx)

        print(f"DEBUG: Found {len(text_tag_rows)} rows with non-numeric tags (action={tag_action})")

        if tag_action == 'skip':
            # Current default: exclude from main df, put in Skipped sheet
            skipped_df = text_tag_rows
            df = df[~non_numeric_mask].copy()

        elif tag_action == 'keep':
            # Use the text tag as-is — leave rows in df, no modification needed
            df = df.copy()  # all rows stay

        elif tag_action == 'extract':
            # Extract the first number from the text tag; fall back to text if none found
            def _extract_tag(x):
                if _is_text_tag(x):
                    m = re.search(r'\d+', str(x))
                    return m.group(0) if m else x
                return x
            df[tag_col] = df[tag_col].apply(_extract_tag)
            df = df.copy()

        else:
            # Unknown action: default to skip
            skipped_df = text_tag_rows
            df = df[~non_numeric_mask].copy()
    else:
        # Just apply the notes mask if no tag column
        df = df[mask].copy()
    
    # Forward fill Tag values (rows with NaN Tag belong to previous Tag)
    # Handle different column names: "Tag", "Tag/Unit", "Tag/unit", or "unit" (case-insensitive)
    tag_col = None
    for col in df.columns:
        col_upper = str(col).upper()
        if col_upper in ['TAG', 'TAG/UNIT', 'UNIT']:
            tag_col = col
            break
    
    if tag_col:
        df[tag_col] = df[tag_col].ffill()
        # Only convert to numeric when we're not deliberately keeping text tags.
        # pd.to_numeric() would silently destroy tags like 'th105' → NaN.
        if tag_action not in ('keep', 'extract'):
            try:
                df[tag_col] = pd.to_numeric(df[tag_col])
            except (ValueError, TypeError):
                pass
    
    # Remove rows where essential columns are all NaN
    # Handle both 'Height' and 'Length' column names
    height_col = 'Height' if 'Height' in df.columns else ('Length' if 'Length' in df.columns else None)
    essential_cols = ['Width', 'Fabric']
    if height_col:
        essential_cols.append(height_col)
    if all(col in df.columns for col in essential_cols):
        df = df[df[essential_cols].notna().any(axis=1)].copy()
    
    # Debug: Check column names
    print(f"Available columns: {list(df.columns)}")
    print(f"Rows after filtering: {len(df)}")
    
    # Build cleaned dataframe
    cleaned_data = []
    
    # Determine tag column name - handle multiple possible names (case-insensitive)
    tag_col = None
    for col in df.columns:
        col_upper = str(col).upper()
        if col_upper in ['TAG', 'TAG/UNIT', 'UNIT']:
            tag_col = col
            break
    
    skipped_count = 0
    skipped_reasons = {'no_fabric': 0, 'no_width': 0, 'no_height': 0}
    error_row_indices = []  # track which output rows need red highlighting
    
    for idx, row in df.iterrows():
        tag = row.get(tag_col, '') if tag_col else ''
        fabric_raw = row.get('Fabric', '')
        
        # Skip if essential data is missing
        if pd.isna(fabric_raw):
            skipped_reasons['no_fabric'] += 1
            skipped_count += 1
            continue
        
        fabric = normalize_fabric(str(fabric_raw).strip())
        
        # Clean and parse width (handle fractions including Unicode: ¾, ¼, ⅛, etc.)
        width_raw = row.get('Width')
        if not pd.isna(width_raw):
            width = parse_fraction_string(width_raw)
        else:
            width = None
        
        # Handle both 'Height' and 'Length' column names (handle fractions including Unicode)
        height_raw = row.get('Height') if 'Height' in row.index else row.get('Length', None)
        if not pd.isna(height_raw):
            height = parse_fraction_string(height_raw)
        else:
            height = None
        
        # If width or height couldn't be parsed, include as an empty (red-flagged) row
        has_parse_error = False
        if width is None and width_raw is not None and not pd.isna(width_raw):
            skipped_reasons['no_width'] += 1
            has_parse_error = True
        if height is None and height_raw is not None and not pd.isna(height_raw):
            skipped_reasons['no_height'] += 1
            has_parse_error = True
        
        # A row with NO numeric data at all (both raw values missing) still gets skipped silently
        if width is None and height is None and not has_parse_error:
            skipped_count += 1
            continue
        
        # Handle both 'Control' and 'Controls' column names
        control = str(row.get('Control', row.get('Controls', ''))).strip().upper()
        # Handle both 'Deducts' and 'Deducts ' (with trailing space)
        deducts = row.get('Deducts ', '') if 'Deducts ' in row.index else row.get('Deducts', '')
        # Check if Unnamed: 12 column exists (some files don't have it)
        deduct_value = row.get('Unnamed: 12', None) if 'Unnamed: 12' in df.columns else None
        roll = row.get('Roll', '')
        chain = row.get('Chain ', None)
        
        # Determine color number based on fabric type
        # Handle Bed, Liv, and other fabric types (Kitchen, Studio, Den, etc.)
        fabric_upper = fabric.upper()
        color_number = None
        
        # Check fabric_colors dict first (for custom mappings)
        if fabric_colors and fabric in fabric_colors:
            color_number = fabric_colors[fabric]
        elif fabric_colors:
            # Try case-insensitive match
            for fab_key, fab_color in fabric_colors.items():
                if fab_key.upper() == fabric_upper:
                    color_number = fab_color
                    break
        
        # If not found in fabric_colors, use standard mappings
        if color_number is None:
            if 'BED' in fabric_upper:
                color_number = bed_color if bed_color else 'BED_COLOR_NEEDED'
            elif 'LIV' in fabric_upper:
                color_number = liv_color if liv_color else 'LIV_COLOR_NEEDED'
            else:
                # For other fabric types (Kitchen, Studio, Den, etc.), use placeholder
                # User can manually update these later
                color_number = f'{fabric.upper()}_COLOR_NEEDED'
        
        # Create Room (Tag-Fabric)
        # Handle tag: can be NaN, empty string, or numeric
        if pd.isna(tag) or tag == '' or str(tag).strip() == '':
            tag_str = 'UNKNOWN'
        else:
            try:
                tag_str = str(int(float(tag)))  # Convert to float first to handle decimals, then int
            except (ValueError, TypeError):
                # If it cannot be converted to float/int, it's a text tag like 'th105'
                tag_str = str(tag).strip()
        room = f"{tag_str}-{fabric}"
        
        # Map Control to ControlSide (accepts L/R/Left/Right, case-insensitive)
        control_side_map = {'L': 'LEFT', 'R': 'RIGHT', 'LEFT': 'LEFT', 'RIGHT': 'RIGHT'}
        control_side = control_side_map.get(control, None)
        
        # Parse deduction - SIMPLE LOGIC
        # When notes contain "D=1":
        #   - Deducts = "D"  → ExtraFabricDeduction = 1.0 (total)
        #   - Deducts = "Dl" → ExtraFabricDeduction = 0.5 (half of D)
        #   - Deducts = "Dr" → ExtraFabricDeduction = 0.5 (half of D)
        extra_fabric_deduction = None
        deduction_code = None  # Track the deduction code (D, Dl, or Dr) for special instructions
        deduction_per_side_value = None  # Track per-side value for "D" deductions (for display in special instructions)
        
        if not pd.isna(deducts) and str(deducts).strip() != '':
            deducts_str = str(deducts).strip()
            
            # Extract code part (D, Dl, or Dr)
            code_match = re.match(r'^(D[LR]?)', deducts_str, re.IGNORECASE)
            if code_match:
                code_part = code_match.group(1).upper()
                deduction_code = code_part  # Store for special instructions
                
                # Get the deduction value using parse_deduction_value
                extra_fabric_deduction = parse_deduction_value(deducts_str, deduct_value, "", deduction_values)
                
                # Debug output for first few rows
                if idx < 3:
                    print(f"DEBUG Row {idx}: Deducts='{deducts_str}', code_part='{code_part}', extra_fabric_deduction={extra_fabric_deduction}")
                
                # Extract per-side value for "D" deductions (for display in special instructions only)
                if code_part == 'D' and extra_fabric_deduction is not None:
                    # Per-side value is half of total for display (e.g., D=0.5 → show "1/4" in special instructions)
                    deduction_per_side_value = extra_fabric_deduction / 2
                    if idx < 3:
                        print(f"DEBUG Row {idx}: deduction_per_side_value={deduction_per_side_value}")
        
        # ControlLength - set for all rows that have Chain value (present in raw file)
        # parse_chain_value handles both numeric (60) and string ("60", '60"') forms
        parsed_chain = parse_chain_value(chain)
        control_length = format_control_length(chain) if parsed_chain is not None else None
        
        # ReverseRoll - if Roll="Rev" then "Yes", otherwise "No" (not None)
        reverse_roll = determine_reverse_roll(roll)
        if reverse_roll is None:
            reverse_roll = "No"
        
        # Special Instructions
        # For "D" deductions, use per-side value for display (e.g., show "1/2"" instead of "1"")
        display_deduction_value = deduction_per_side_value if (deduction_code == 'D' and deduction_per_side_value is not None) else extra_fabric_deduction
        special_instructions = create_special_instructions(
            reverse_roll == "Yes",
            control_side or "",
            display_deduction_value,
            deduction_code
        )
        
        # Set ExtraFabricDeduction to 0 if None/NaN (no deduction)
        if extra_fabric_deduction is None:
            extra_fabric_deduction = 0.0
        
        row_data = {
            'Color Number': color_number,
            'Width': width,
            'Height': int(height) if height is not None else None,
            'ExtraFabricDeduction': extra_fabric_deduction,
            'ReverseRoll': reverse_roll,
            'ControlSide': control_side,
            'ControlLength': control_length,
            'Room': room,
            'SpecialInstructions': special_instructions
        }
        if has_parse_error:
            error_row_indices.append(len(cleaned_data))  # index before append
        cleaned_data.append(row_data)
    
    df_cleaned = pd.DataFrame(cleaned_data)
    
    # Clean up trailing spaces in Room and ControlSide
    if 'Room' in df_cleaned.columns:
        df_cleaned['Room'] = df_cleaned['Room'].str.replace(r'\s+$', '', regex=True)
    if 'ControlSide' in df_cleaned.columns:
        df_cleaned['ControlSide'] = df_cleaned['ControlSide'].str.replace(r'\s+$', '', regex=True)
    
    # Save to file if output_file specified
    if output_file:
        output_path = output_file
    else:
        # Auto-generate output filename
        input_path = Path(input_file)
        output_path = input_path.parent / f"{input_path.stem}-cleaned.xlsx"
    
    # Save with wrap text formatting for SpecialInstructions column and highlighting for width > 144
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_cleaned.to_excel(writer, index=False, sheet_name='Cleaned')
        
        # Get the worksheet for formatting
        worksheet = writer.sheets['Cleaned']
        
        # Yellow fill for wide-width highlighting; red fill for parse-error rows
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF4C4C', end_color='FF4C4C', fill_type='solid')
        
        # Convert error_row_indices (0-based DataFrame rows) to Excel row numbers (2-based)
        error_excel_rows = {i + 2 for i in error_row_indices}
        
        # Find column indices
        width_col_idx = df_cleaned.columns.get_loc('Width') + 1 if 'Width' in df_cleaned.columns else None
        special_instructions_col_idx = df_cleaned.columns.get_loc('SpecialInstructions') + 1 if 'SpecialInstructions' in df_cleaned.columns else None
        
        # Apply formatting to each row
        num_cols = len(df_cleaned.columns)
        for row_num in range(2, len(df_cleaned) + 2):  # Start from row 2 (row 1 is header)
            # Red-highlight entire row for parse errors (bad width/height)
            if row_num in error_excel_rows:
                for col_idx in range(1, num_cols + 1):
                    worksheet.cell(row=row_num, column=col_idx).fill = red_fill
            else:
                # Highlight width cell if value > 144 (only for non-error rows)
                if width_col_idx:
                    width_value = df_cleaned.iloc[row_num - 2]['Width']
                    if pd.notna(width_value) and width_value > 144:
                        width_col_letter = openpyxl.utils.get_column_letter(width_col_idx)
                        worksheet[f'{width_col_letter}{row_num}'].fill = yellow_fill
            
            # Enable wrap text for SpecialInstructions column
            if special_instructions_col_idx:
                col_letter = openpyxl.utils.get_column_letter(special_instructions_col_idx)
                cell = worksheet[f'{col_letter}{row_num}']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Enable wrap text for SpecialInstructions header
        if special_instructions_col_idx:
            header_col_letter = openpyxl.utils.get_column_letter(special_instructions_col_idx)
            header_cell = worksheet[f'{header_col_letter}1']
            header_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        
        # Write Skipped rows to a separate sheet if any exist
        if not skipped_df.empty:
            skipped_df.to_excel(writer, index=False, sheet_name='Skipped')
            print(f"Added 'Skipped' sheet with {len(skipped_df)} rows")
    
    print(f"Cleaned file saved to: {output_path}")
    
    if skipped_count > 0:
        print(f"\nDEBUG: Skipped {skipped_count} rows during processing:")
        for reason, count in skipped_reasons.items():
            if count > 0:
                print(f"  - {reason}: {count}")
    
    return df_cleaned


def detect_color_numbers(df_cleaned: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    """Try to detect color numbers from cleaned data (if available)"""
    bed_color = None
    liv_color = None
    
    if 'Color Number' in df_cleaned.columns and 'Room' in df_cleaned.columns:
        bed_rows = df_cleaned[df_cleaned['Room'].str.contains('-Bed', na=False, case=False)]
        liv_rows = df_cleaned[df_cleaned['Room'].str.contains('-Liv', na=False, case=False)]
        
        if len(bed_rows) > 0:
            bed_colors = bed_rows['Color Number'].unique()
            if len(bed_colors) == 1:
                bed_color = bed_colors[0]
        
        if len(liv_rows) > 0:
            liv_colors = liv_rows['Color Number'].unique()
            if len(liv_colors) == 1:
                liv_color = liv_colors[0]
    
    return bed_color, liv_color


if __name__ == "__main__":
    import sys
    
    # Simple command-line interface (backward compatible)
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else None
        bed_color = sys.argv[3] if len(sys.argv) > 3 else None
        liv_color = sys.argv[4] if len(sys.argv) > 4 else None
        deduction_cell = sys.argv[5] if len(sys.argv) > 5 else 'I6'  # Default to I6
    else:
        # Default: clean the example file
        input_file = "Symington Mid Rise L1 to L3 .xlsx"
        output_file = None
        bed_color = "YUNOWH"  # From the cleaned example
        liv_color = "PWS3WHIT"  # From the cleaned example
        deduction_cell = 'I6'  # Default cell for D value
    
    fabric_colors = None
    
    print("=" * 60)
    print("Excel Cleaning Script")
    print("=" * 60)
    print(f"Reading D value from cell: {deduction_cell}")
    print("(You can change this by passing a 5th argument, e.g., 'I6')")
    print()
    
    if not bed_color or not liv_color:
        print("\nWARNING: Color numbers not provided.")
        print("   Bed color:", bed_color or "NOT SET")
        print("   Liv color:", liv_color or "NOT SET")
        print("   You can provide them as command line arguments:")
        print("   python clean_excel.py <input_file> [output_file] [bed_color] [liv_color] [deduction_cell]")
        print()
    
    df_cleaned = clean_excel_file(
        input_file=input_file,
        output_file=output_file,
        bed_color=bed_color,
        liv_color=liv_color,
        fabric_colors=fabric_colors,
        header_row=8,  # Default header row
        deduction_cell=deduction_cell
    )
    
    print("\n" + "=" * 60)
    print("Cleaning Complete!")
    print("=" * 60)
    print(f"\nCleaned {len(df_cleaned)} rows")
    print(f"\nFirst few rows:")
    print(df_cleaned.head(10).to_string())
    
    # Check for missing color numbers
    if 'Color Number' in df_cleaned.columns:
        missing_colors = df_cleaned[df_cleaned['Color Number'].str.contains('NEEDED', na=False)]
        if len(missing_colors) > 0:
            print(f"\nWARNING: {len(missing_colors)} rows have missing color numbers")
            unique_fabrics = missing_colors['Room'].str.split('-').str[1].unique()
            print(f"   Fabric types needing colors: {', '.join(unique_fabrics)}")
            print("   You can provide colors via fabric_colors parameter or update manually")

