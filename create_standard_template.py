"""
Create a standard Excel template for window shade data entry
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Window Shades'

# Define styles
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
metadata_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
metadata_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# NOTES SECTION - Top rows (Column A, always visible)
notes_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
notes_font = Font(bold=True, size=10, italic=True)

ws['A1'] = 'NOTES:'
ws['A1'].font = Font(bold=True, size=11, color='FF0000')
ws['A1'].fill = notes_fill

ws['A2'] = '• Deducts: Dl = Left side, Dr = Right side, D = Both sides'
ws['A2'].font = notes_font
ws['A2'].fill = notes_fill

ws['A3'] = '• Tag/Unit can be blank - blank rows belong to previous unit'
ws['A3'].font = notes_font
ws['A3'].fill = notes_fill

ws['A4'] = '• D value in I7 is TOTAL for both sides (Dl and Dr are each D/2)'
ws['A4'].font = notes_font
ws['A4'].fill = notes_fill

# PROJECT INFO & COLORS - Column F (rows 1-7)
ws['F1'] = 'Project Name:'
ws['G1'] = 'Sample Project - Building Name'
ws['F1'].font = metadata_font
ws['F1'].fill = metadata_fill

ws['F2'] = 'Date:'
ws['G2'] = '2026-01-21'
ws['F2'].font = metadata_font
ws['F2'].fill = metadata_fill

ws['F3'] = 'Bed/Bedroom ='
ws['G3'] = 'YUNOWH'
ws['F3'].font = metadata_font
ws['F3'].fill = metadata_fill
ws['G3'].font = Font(bold=True, color='FF0000')

ws['F4'] = 'Living/Liv ='
ws['G4'] = 'PWS3WHIT'
ws['F4'].font = metadata_font
ws['F4'].fill = metadata_fill
ws['G4'].font = Font(bold=True, color='FF0000')

ws['F5'] = 'Studio ='
ws['G5'] = 'STUDIO123'
ws['F5'].font = metadata_font
ws['F5'].fill = metadata_fill
ws['G5'].font = Font(bold=True, color='0000FF')

ws['F6'] = 'Kitchen ='
ws['G6'] = 'KITCHEN456'
ws['F6'].font = metadata_font
ws['F6'].fill = metadata_fill
ws['G6'].font = Font(bold=True, color='0000FF')

# D Value - Cell I7 (CRITICAL - Yellow highlight)
ws['I7'] = 'D = 1/2'
ws['I7'].font = Font(bold=True, size=12, color='FF0000')
ws['I7'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ws['I7'].alignment = Alignment(horizontal='center', vertical='center')

# Add comment to I7
ws['I7'].comment = openpyxl.comments.Comment(
    'IMPORTANT: D value for deductions.\nD = total for both sides.\nDl and Dr are each D/2.',
    'Template'
)

# Row 8: Empty separator

# Row 9: Headers (0-indexed row 8)
headers = ['Tag/Unit', 'Q', 'Product', 'Roll', 'Width', 'Height', 'Chain ', 'Fabric', 'Control', 'Deducts ']
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=9, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Sample data rows (starting row 10)
sample_data = [
    [120, 1, 'Manual Shade', 'Rev', 32, 86, 72, 'Bed', 'L', ''],
    ['', 1, 'Manual Shade', 'Rev', 62.25, 86, 72, 'Bed', 'R', ''],
    ['', 1, 'Manual Shade', 'Rev', 42.25, 122, 72, 'Liv', 'L', 'Dl'],
    ['', 1, 'Manual Shade', 'Rev', 52, 122, 72, 'Liv', 'R', 'Dr'],
    [121, 1, 'Manual Shade', '', 48, 96, 60, 'Bed', 'L', ''],
    ['', 1, 'Manual Shade', '', 55.5, 96, 60, 'Bed', 'R', ''],
    [122, 1, 'Manual Shade', 'Rev', 36, 108, 72, 'Liv', 'L', 'D'],
    [223, 1, 'Manual Shade', '', 44, 84, 48, 'Studio', 'L', ''],
    [224, 1, 'Manual Shade', '', 50, 90, 60, 'Kitchen', 'R', 'Dl'],
]

for row_idx, row_data in enumerate(sample_data, start=10):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border
        if col_idx in [5, 6, 7]:  # Numeric columns (Width, Height, Chain)
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='center')

# Set column widths (adjusted for notes in column A)
column_widths = [50, 5, 15, 6, 10, 10, 10, 10, 10, 10]  # Column A wider for notes, removed Drawing Ref., Mount, and Unnamed: 12
for col_idx, width in enumerate(column_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Freeze panes (freeze header row)
ws.freeze_panes = 'A10'

# Add instruction sheet
ws2 = wb.create_sheet('Instructions')
ws2['A1'] = 'STANDARD RAW FILE TEMPLATE - INSTRUCTIONS'
ws2['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws2['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

instructions = [
    '',
    'HOW TO USE THIS TEMPLATE:',
    '',
    '1. LAYOUT OVERVIEW:',
    '   • Column A (Rows 1-4): NOTES - Always visible reference information',
    '   • Column F-G (Rows 1-7): PROJECT INFO & COLOR CODES',
    '   • Cell I7: D VALUE (Yellow highlight) - CRITICAL!',
    '   • Row 9: HEADERS - Do not modify',
    '   • Row 10+: YOUR DATA',
    '',
    '2. UPDATE PROJECT INFO (Column F-G):',
    '   • G1: Update Project Name',
    '   • G2: Update Date',
    '   • G3: IMPORTANT - Update Bed/Bedroom color code (Red text)',
    '   • G4: IMPORTANT - Update Living/Liv color code (Red text)',
    '   • G5: Update Studio color code (Blue text, optional)',
    '   • G6: Update Kitchen color code (Blue text, optional)',
    '',
    '3. UPDATE D VALUE (Cell I7):',
    '   • CRITICAL: Update D value (e.g., "D = 1/2" or "D = 1")',
    '   • Cell I7 is highlighted in YELLOW',
    '   • D = total deduction for BOTH sides',
    '   • Dl and Dr are automatically D/2 each',
    '',
    '4. DATA ENTRY (Row 10+):',
    '   • Row 9 contains headers - DO NOT MODIFY HEADER NAMES',
    '   • Start entering data from Row 10',
    '   • Tag: Room/unit number (can leave blank for continuation rows)',
    '   • Width: Window width in inches (required)',
    '   • Height: Window height in inches (required)',
    '   • Fabric: Bed, Liv, Studio, Kitchen, Den, Bath (required)',
    '   • Control: L (Left) or R (Right) (required)',
    '   • Roll: Enter "Rev" for reverse roll, leave blank otherwise',
    '   • Chain: Chain length value (e.g., 72, 48, 60)',
    '   • Deducts: Dl (left), Dr (right), or D (both sides)',
    '   • Deducts can include value: "Dl=1/2", "Dr=1/2", "D=1" or just "Dl", "Dr", "D"',
    '',
    '5. IMPORTANT NOTES:',
    '   • Headers MUST be at Row 9 (0-indexed: row 8)',
    '   • Column names must match exactly (including spaces in "Chain " and "Deducts ")',
    '   • Tag can be blank - blank rows belong to previous Tag number',
    '   • D value in I7 is the TOTAL for both sides (Dl and Dr are each D/2)',
    '   • Notes in Column A are always visible for quick reference',
    '',
    '6. DEDUCTION LOGIC:',
    '   • If I7 contains "D = 1/2":',
    '     - Dl = 0.25 (left side, half of D)',
    '     - Dr = 0.25 (right side, half of D)',
    '     - D = 0.5 (both sides, total)',
    '   • If I7 contains "D = 1":',
    '     - Dl = 0.5 (left side, half of D)',
    '     - Dr = 0.5 (right side, half of D)',
    '     - D = 1.0 (both sides, total)',
    '',
    '7. EXAMPLE DATA:',
    '   • Tag 120 has 4 rows (first row has Tag=120, next 3 are blank)',
    '   • All 4 rows belong to room 120',
    '   • Some have deductions (Dl, Dr, D), some don\'t',
    '   • Some have reverse roll (Rev), some don\'t',
    '',
    '8. RUNNING THE CLEANING SCRIPT:',
    '   python clean_excel.py "your_file.xlsx"',
    '',
    '   Or with custom parameters:',
    '   python clean_excel.py "your_file.xlsx" "output.xlsx" "BED_COLOR" "LIV_COLOR" "I7"',
    '',
    '9. OUTPUT:',
    '   • Creates file: your_file-cleaned.xlsx',
    '   • Contains 9 columns:',
    '     - Color Number',
    '     - Width',
    '     - Height',
    '     - ExtraFabricDeduction',
    '     - ReverseRoll',
    '     - ControlSide',
    '     - ControlLength',
    '     - Room',
    '     - SpecialInstructions',
    '',
    '10. TIPS:',
    '   • Delete sample data rows (10-18) before entering your data',
    '   • Keep the header row (row 9) intact',
    '   • Update color codes and D value before running script',
    '   • Use "Rev" exactly for reverse roll (case-insensitive)',
    '   • Leave Tag blank for continuation rows',
    '   • Notes in Column A provide quick reference while entering data',
]

for idx, instruction in enumerate(instructions, start=1):
    cell = ws2[f'A{idx+1}']
    cell.value = instruction
    if instruction.startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.', '10.')):
        cell.font = Font(bold=True, size=11)
    elif instruction.startswith('   •'):
        cell.font = Font(size=10)
    elif instruction.startswith('     -'):
        cell.font = Font(size=9, italic=True)

ws2.column_dimensions['A'].width = 100

# Save file
wb.save('STANDARD_TEMPLATE.xlsx')
print('✅ Created STANDARD_TEMPLATE.xlsx')
print('')
print('📋 Template Details:')
print('   • Sheet 1: Window Shades (with sample data)')
print('   • Sheet 2: Instructions (detailed guide)')
print('   • Column A (Rows 1-4): Notes - Always visible!')
print('   • Column F-G (Rows 1-7): Project info & color codes')
print('   • Cell I7: D value (highlighted in YELLOW)')
print('   • Row 9: Headers (do not modify)')
print('   • Rows 10-18: Sample data (delete before use)')
print('')
print('🎯 Next Steps:')
print('   1. Open STANDARD_TEMPLATE.xlsx')
print('   2. Update color codes (G3, G4, G5, G6)')
print('   3. Update D value (I7)')
print('   4. Replace sample data with your data')
print('   5. Run: python clean_excel.py "your_file.xlsx"')
